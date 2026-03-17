from django.shortcuts import render
from django.http import HttpResponse
from django.contrib.auth.decorators import login_required
from openpyxl import Workbook
from django.utils.timezone import localtime
import json
from .models import LoginHistory, NavigationEvent, LogoutHistory

@login_required
def audit_dashboard(request):
    user_filter = request.GET.get('user', None)
    date_filter = request.GET.get('date', None)

    login_history = LoginHistory.objects.all()
    navigation_events = NavigationEvent.objects.all()
    logout_history = LogoutHistory.objects.all()

    if user_filter:
        login_history = login_history.filter(user__username__icontains=user_filter)
        navigation_events = navigation_events.filter(user__username__icontains=user_filter)
        logout_history = logout_history.filter(user__username__icontains=user_filter)

    if date_filter:
        login_history = login_history.filter(login_time__date=date_filter)
        navigation_events = navigation_events.filter(timestamp__date=date_filter)
        logout_history = logout_history.filter(logout_time__date=date_filter)

    context = {
        'login_history': login_history,
        'navigation_events': navigation_events,
        'logout_history': logout_history,
    }
    return render(request, 'audit/dashboard.html', context)

@login_required
def export_reports(request):
    workbook = Workbook()

    # Login History Sheet
    login_sheet = workbook.active
    login_sheet.title = "Login History"
    login_sheet.append(["User", "Login Time", "Channel"])
    for record in LoginHistory.objects.all():
        login_sheet.append([
            record.user.username,
            localtime(record.login_time).replace(tzinfo=None),
            record.channel
        ])

    # Navigation Events Sheet
    nav_sheet = workbook.create_sheet(title="Navigation Events")
    nav_sheet.append(["User", "URL", "Method", "Payload", "Parameters", "Headers", "Coming From", "Timestamp"])
    for record in NavigationEvent.objects.all():
        nav_sheet.append([
            record.user.username if record.user else "Anonymous",
            record.url,
            record.method,
            json.dumps(record.payload),
            record.parameters,
            json.dumps(record.headers),
            record.coming_from,
            localtime(record.timestamp).replace(tzinfo=None)
        ])

    # Logout History Sheet
    logout_sheet = workbook.create_sheet(title="Logout History")
    logout_sheet.append(["User", "Logout Time"])
    for record in LogoutHistory.objects.all():
        logout_sheet.append([
            record.user.username,
            localtime(record.logout_time).replace(tzinfo=None)
        ])

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="audit_reports.xlsx"'
    workbook.save(response)
    return response
