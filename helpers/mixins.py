from django.shortcuts import get_object_or_404
from django.contrib.contenttypes.models import ContentType
from django.core.exceptions import ValidationError
import io, datetime
from django.http import HttpResponse
from django.db import models as django_models
import openpyxl
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter

from helpers.utilities import get_exportable_fields, resolve_exportable_value

"""
ExportListMixin
---------------
Drop-in mixin for Django ListViews.

Features
--------
* GET ?export=excel  → streams an .xlsx file of the current queryset
* Auto-detects all concrete model fields
* id, uid, and raw FK _id columns are always excluded
* Choice fields (status, priority, etc.) use get_FOO_display() labels
* Alternating row colours  (navy / steel-blue)
* Auto column width, frozen header row
* DateTime / Date formatted as YYYY-MM-DD
* M2M fields hidden but values listed as "1. A, 2. B"
* is_active, created, updated are always the last columns
* @can_export decorated properties/methods are included automatically
  - position="before" → prepended before regular fields
  - position="after"  → inserted after regular fields, before trailing fields

Usage
-----
    class ProjectListView(LoginRequiredMixin, ExportListMixin, SafeListView):
        model = Project
        template_name = "projects/index.html"
        context_object_name = "projects"
"""

OPENPYXL_AVAILABLE = True

# ── Colour palette ──────────────────────────────────────────────────────────
HEADER_BG  = "1E3A5F"
HEADER_FG  = "FFFFFF"
ROW_A_BG   = "D6E4F0"
ROW_B_BG   = "EBF5FB"
BORDER_COL = "A9CCE3"

# Fields always excluded regardless of model
ALWAYS_EXCLUDE = {
    "id", "uid", "password", "is_deleted"
}

# Fields that should always appear as the last columns (in this order)
TRAILING_FIELDS = ["is_active", "created", "updated"]

# Any attname ending with these is a raw FK column (company_id, user_id …)
FK_SUFFIXES = ("_id",)

# Column kind sentinels
_COL_REGULAR = "regular"
_COL_M2M     = "m2m"
_COL_CUSTOM  = "custom"   # @can_export property / method


def _thin_border():
    side = Side(style="thin", color=BORDER_COL)
    return Border(left=side, right=side, top=side, bottom=side)


def _header_fill():
    return PatternFill("solid", fgColor=HEADER_BG)


def _row_fill(even: bool):
    return PatternFill("solid", fgColor=ROW_A_BG if even else ROW_B_BG)


# ── Field introspection ──────────────────────────────────────────────────────

def _should_exclude(attname: str) -> bool:
    if attname in ALWAYS_EXCLUDE:
        return True
    if attname.endswith(FK_SUFFIXES):
        return True
    return False


def _get_field_info(model):
    """
    Returns:
        regular_fields – list of (attname, verbose_name, field_obj)
        m2m_fields     – list of (attname, verbose_name)

    id, uid, and raw FK _id columns are skipped automatically.
    is_active, created, updated are always moved to the end (in that order).
    """
    opts = model._meta
    regular, m2m = [], []

    for f in opts.get_fields():
        if f.is_relation and (f.one_to_many or (f.many_to_one and not f.concrete)):
            continue
        if isinstance(f, django_models.ManyToManyField):
            m2m.append((f.name, str(f.verbose_name)))
            continue
        if hasattr(f, "attname"):
            if _should_exclude(f.attname):
                continue
            regular.append((f.attname, str(f.verbose_name), f))

    trailing_lookup = {attname: i for i, attname in enumerate(TRAILING_FIELDS)}
    main_fields     = [t for t in regular if t[0] not in trailing_lookup]
    trailing_fields = sorted(
        [t for t in regular if t[0] in trailing_lookup],
        key=lambda t: trailing_lookup[t[0]],
    )
    return main_fields + trailing_fields, m2m


def _cell_value(value, field, obj=None):
    if obj is not None and field is not None:
        choices = getattr(field, "choices", None)
        if choices:
            display_fn = f"get_{field.name}_display"
            if hasattr(obj, display_fn):
                value = getattr(obj, display_fn)()

    if value is None:
        return ""
    if isinstance(value, bool):
        return "Yes" if value else "No"
    if isinstance(value, datetime.datetime):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, datetime.date):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, (int, float)):
        return value
    return str(value)


def _m2m_value(obj, attname):
    try:
        manager = getattr(obj, attname)
        items = list(manager.all())
        if not items:
            return ""
        return ", ".join(f"{i+1}. {item}" for i, item in enumerate(items))
    except Exception:
        return ""


# ── Column assembly ───────────────────────────────────────────────────────────

def _build_visible_cols(regular_fields, m2m_fields, exportable_fields, exclude: set):
    """
    Assemble the full ordered column list for the spreadsheet.

    Final column order:
        1. @can_export(position="before")   ← prepended
        2. regular model fields             ← main body
        3. @can_export(position="after")    ← appended after body
        4. trailing fields                  ← is_active, created, updated
        5. M2M fields                       ← always hidden

    Each entry is a dict:
        { "attname", "label", "kind", "field" }
    """
    trailing_lookup  = {a: i for i, a in enumerate(TRAILING_FIELDS)}

    main_regular     = [t for t in regular_fields if t[0] not in trailing_lookup and t[0] not in exclude]
    trailing_regular = [t for t in regular_fields if t[0] in trailing_lookup     and t[0] not in exclude]

    before_custom = [f for f in exportable_fields if f["position"] == "before" and f["attname"] not in exclude]
    after_custom  = [f for f in exportable_fields if f["position"] == "after"  and f["attname"] not in exclude]

    m2m_cols = [
        {"attname": a, "label": v.title(), "kind": _COL_M2M, "field": None}
        for a, v in m2m_fields if a not in exclude
    ]

    cols = []

    for f in before_custom:
        cols.append({"attname": f["attname"], "label": f["label"], "kind": _COL_CUSTOM, "field": None})

    for attname, vname, field in main_regular:
        cols.append({"attname": attname, "label": vname.title(), "kind": _COL_REGULAR, "field": field})

    for f in after_custom:
        cols.append({"attname": f["attname"], "label": f["label"], "kind": _COL_CUSTOM, "field": None})

    for attname, vname, field in trailing_regular:
        cols.append({"attname": attname, "label": vname.title(), "kind": _COL_REGULAR, "field": field})

    cols.extend(m2m_cols)
    return cols


# ── Main mixin ───────────────────────────────────────────────────────────────

class ExportListMixin:
    """
    Add to any ListView to get ?export=excel support.

    Class-level overrides:
        export_filename (str)          – base name for downloaded file
        export_exclude_fields (list)   – extra field attnames to skip
        export_sheet_name (str)        – worksheet tab name
    """

    export_filename: str = ""
    export_exclude_fields: list = []
    export_sheet_name: str = "Export"

    def get(self, request, *args, **kwargs):
        if request.GET.get("export") == "excel":
            return self._export_excel(request)
        return super().get(request, *args, **kwargs)

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx["export_excel_url"] = self._build_export_url()

        model        = self.get_queryset().model
        regular, m2m = _get_field_info(model)
        exportable   = get_exportable_fields(model)
        exclude      = set(self.export_exclude_fields)
        visible_cols = _build_visible_cols(regular, m2m, exportable, exclude)

        ctx["table_headers"] = [
            col["label"] for col in visible_cols if col["kind"] != _COL_M2M
        ]
        return ctx

    def _build_export_url(self):
        params = self.request.GET.copy()
        params["export"] = "excel"
        return f"?{params.urlencode()}"

    def _export_excel(self, request):
        qs    = self.get_queryset()
        model = qs.model

        regular_fields, m2m_fields = _get_field_info(model)
        exportable_fields          = get_exportable_fields(model)
        exclude                    = set(self.export_exclude_fields)
        visible_cols               = _build_visible_cols(
            regular_fields, m2m_fields, exportable_fields, exclude
        )

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = self.export_sheet_name

        border       = _thin_border()
        header_font  = Font(bold=True, color=HEADER_FG, name="Calibri", size=11)
        data_font    = Font(name="Calibri", size=10)
        center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        left_align   = Alignment(horizontal="left",   vertical="center", wrap_text=True)

        # ── header row ────────────────────────────────────────────────────────
        for col_idx, col in enumerate(visible_cols, start=1):
            cell           = ws.cell(row=1, column=col_idx, value=col["label"])
            cell.fill      = _header_fill()
            cell.font      = header_font
            cell.border    = border
            cell.alignment = center_align

        # ── data rows ─────────────────────────────────────────────────────────
        for row_num, obj in enumerate(qs, start=2):
            fill = _row_fill(row_num % 2 == 0)

            for col_idx, col in enumerate(visible_cols, start=1):
                kind    = col["kind"]
                attname = col["attname"]

                if kind == _COL_M2M:
                    raw = _m2m_value(obj, attname)
                elif kind == _COL_CUSTOM:
                    raw = resolve_exportable_value(obj, attname)
                else:
                    raw = getattr(obj, attname, "")
                    raw = _cell_value(raw, col["field"], obj=obj)

                cell           = ws.cell(row=row_num, column=col_idx, value=raw)
                cell.fill      = fill
                cell.font      = data_font
                cell.border    = border
                cell.alignment = left_align

        # ── column widths & hide M2M ──────────────────────────────────────────
        for col_idx, col in enumerate(visible_cols, start=1):
            col_letter = get_column_letter(col_idx)
            if col["kind"] == _COL_M2M:
                ws.column_dimensions[col_letter].hidden = True
                ws.column_dimensions[col_letter].width  = 0
            else:
                max_len = len(str(ws.cell(row=1, column=col_idx).value or ""))
                for r in range(2, min(ws.max_row + 1, 52)):
                    val = ws.cell(row=r, column=col_idx).value
                    max_len = max(max_len, len(str(val or "")))
                ws.column_dimensions[col_letter].width = min(max_len + 4, 50)

        ws.freeze_panes = "A2"

        # ── stream ────────────────────────────────────────────────────────────
        filename = (
            self.export_filename
            or model._meta.verbose_name_plural.replace(" ", "_")
        )
        filename = f"{filename}_{datetime.date.today()}.xlsx"

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        response = HttpResponse(
            buf.read(),
            content_type=(
                "application/vnd.openxmlformats-officedocument"
                ".spreadsheetml.sheet"
            ),
        )
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        return response


# ── Other existing mixins (unchanged) ────────────────────────────────────────

class CompanyMixin:
    """Automatically sets company field and validates constraints."""
    def form_valid(self, form):
        if self.request.user and self.request.user.company:
            form.instance.company = self.request.user.company
        try:
            form.instance.validate_constraints()
        except ValidationError as e:
            self._add_validation_errors_to_form(form, e)
            return self.form_invalid(form)
        try:
            return super().form_valid(form)
        except Exception:
            return self.form_invalid(form)


class UIDObjectMixin:
    model = None

    def get_object(self):
        uid = self.kwargs.get('uid')
        return get_object_or_404(self.model, uid=uid)


class GenericAttachMixin:
    """Resolves any model instance from content_type + uid."""
    def get_target_object(self):
        content_type_id = self.kwargs.get("content_type_id")
        obj_uid         = self.kwargs.get("obj_uid")
        content_type    = get_object_or_404(ContentType, id=content_type_id)
        model_class     = content_type.model_class()
        return get_object_or_404(model_class, uid=obj_uid)


class UserTrackingMixin:
    """Automatically sets created_by to the logged-in user on create."""
    def form_valid(self, form):
        if self.request.user and self.request.user.is_authenticated:
            form.instance.created_by = self.request.user
        return super().form_valid(form)